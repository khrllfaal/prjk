#!/usr/bin/env python3
"""
Import the company's Excel workbook (KAS BESAR/KECIL, BANK*, COA,
MASTER CUST & VENDOR, Nama Project, JURNAL UMUM) into the MySQL/MariaDB
database created by backend/mysql/schema.sql (primary backend, used for
Hostinger deployment).

This script is intentionally NOT wired to any hardcoded file path or
committed data — it takes the xlsx path and DB connection as arguments/
env vars so the real financial data never has to touch git.

Usage:
    export MYSQL_HOST=127.0.0.1 MYSQL_DB=accv2 MYSQL_USER=accv2_user MYSQL_PASS=...
    python3 scripts/import_excel_mysql.py /path/to/LAPORAN_KEUANGAN.xlsx

Options:
    --dry-run   parse and print row counts only, no DB writes
    --wipe      truncate the target tables before importing (fresh import)

Safe to re-run: every insert is an upsert (INSERT ... ON DUPLICATE KEY
UPDATE), matching the same "always upsert by id" contract the PHP API
and frontend already use, so running it twice does not duplicate rows.
"""
import argparse
import os
import re
import sys
from datetime import date, datetime

import openpyxl

CASH_SHEETS = {
    "KAS BESAR": "Kas Besar",
    "KAS KECIL": "Kas Kecil",
    "BANK BCA": "Bank BCA",
    "BANK BJB": "Bank BJB",
    "BANK BRI": "Bank BRI",
    "BANK BNI": "Bank BNI",
    "MANDIRI": "Bank MANDIRI",
}


def header_index(row):
    """Map non-empty header cell text -> column index (0-based)."""
    idx = {}
    for i, v in enumerate(row):
        if v:
            idx[str(v).strip()] = i
    return idx


def as_date(v):
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    if isinstance(v, str) and v.strip():
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
            try:
                return datetime.strptime(v.strip(), fmt).date()
            except ValueError:
                pass
    return None


def as_num(v):
    if v is None or v == "":
        return 0.0
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0.0


def load_coa(wb):
    ws = wb["COA"]
    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        kode = r[0]
        if not kode or not re.match(r"^\d+(\.\d+)*$", str(kode)):
            continue
        nama = (r[1] or "").strip()
        level = str(kode).count(".") + 1
        head = str(kode).split(".")[0]
        tipe = {"1": "asset", "2": "liability", "3": "equity", "4": "revenue",
                "5": "cogs", "6": "expense", "7": "other", "8": "tax"}.get(head, "other")
        rows.append({"kode": str(kode), "nama": nama, "level": level, "tipe": tipe})
    return rows


def load_relasi(wb):
    ws = wb["MASTER CUST & VENDOR"]
    customers, vendors = [], []
    ci = vi = 1
    for r in ws.iter_rows(min_row=3, values_only=True):
        nama = r[1] if len(r) > 1 else None
        if not nama or not isinstance(nama, str) or nama.startswith("Note"):
            continue
        rec = {"nama": nama.strip(), "alamat": (r[2] or "").strip() if len(r) > 2 else "",
               "telp": str(r[3]).strip() if len(r) > 3 and r[3] else ""}
        if re.match(r"^Pr-", nama, re.I):
            rec["kode"] = f"CU{ci:03d}"
            customers.append(rec)
            ci += 1
        else:
            rec["kode"] = f"VE{vi:03d}"
            vendors.append(rec)
            vi += 1
    return customers, vendors


PEMBERI_PREFIXES = ["Kemhan", "Bina Marga", "PUPR", "Korem", "Kodim", "Koramil", "BJB", "Dinas", "Diknas"]


def guess_pemberi(nama):
    for p in PEMBERI_PREFIXES:
        if nama.lower().startswith(p.lower()) or f" {p.lower()} " in f" {nama.lower()} ":
            return p
    return ""


def load_projects(wb):
    ws = wb["Nama Project"]
    rows = []
    for r in ws.iter_rows(min_row=3, values_only=True):
        nama = r[1] if len(r) > 1 else None
        if not nama or not isinstance(nama, str):
            continue
        rows.append({"nama": nama.strip(), "kontrak": 0, "rap": 0, "progress": None,
                     "pemberi_proyek": guess_pemberi(nama), "cost_center": 0, "adm_fee": 0})
    return rows


def load_cash_transactions(wb):
    txns = []
    counters = {}
    for sheet, akun_kas in CASH_SHEETS.items():
        if sheet not in wb.sheetnames:
            continue
        ws = wb[sheet]
        rows_iter = ws.iter_rows(min_row=1, values_only=True)
        header = header_index(next(rows_iter))
        is_bank = akun_kas.lower().startswith("bank")
        col_tgl = header.get("Tanggal")
        col_ket = header.get("Keterangan")
        col_relasi = header.get("Customer/Vendor")
        col_project = header.get("Nama Project", header.get("Kategori Produk"))
        col_akun = header.get("Nama Akun")
        col_kategori = header.get("Kategori Akun")
        col_debet = header.get("Debet")
        col_kredit = header.get("Kredit")
        if col_tgl is None or col_akun is None:
            continue
        for r in rows_iter:
            tgl = as_date(r[col_tgl]) if col_tgl < len(r) else None
            akun_lawan = r[col_akun] if col_akun < len(r) else None
            if not tgl or not akun_lawan:
                continue
            debet = as_num(r[col_debet]) if col_debet is not None and col_debet < len(r) else 0.0
            kredit = as_num(r[col_kredit]) if col_kredit is not None and col_kredit < len(r) else 0.0
            if debet == 0 and kredit == 0:
                continue
            is_in = debet > 0
            jenis = ("bank_" if is_bank else "kas_") + ("masuk" if is_in else "keluar")
            pref = ("BI" if is_bank else "CI") if is_in else ("BO" if is_bank else "CO")
            mo = tgl.strftime("%y%m")
            counters[pref] = counters.get(pref, 0) + 1
            relasi = r[col_relasi] if col_relasi is not None and col_relasi < len(r) else None
            relasi = (relasi or "").strip() if isinstance(relasi, str) else ""
            project = r[col_project] if col_project is not None and col_project < len(r) else None
            project = (project or "").strip() if isinstance(project, str) else ""
            ket = r[col_ket] if col_ket is not None and col_ket < len(r) else ""
            kategori = r[col_kategori] if col_kategori is not None and col_kategori < len(r) else ""
            txns.append({
                "tgl": tgl, "ref": f"{pref}-{mo}-{counters[pref]:04d}",
                "akun_kas": akun_kas, "akun_lawan": str(akun_lawan).strip(),
                "project": project, "relasi": relasi,
                "kategori": str(kategori).strip() if kategori else "",
                "ket": str(ket).strip() if ket else "",
                "debet": debet, "kredit": kredit, "jenis": jenis,
            })
    return txns


def load_jurnal(wb):
    if "JURNAL UMUM" not in wb.sheetnames:
        return []
    ws = wb["JURNAL UMUM"]
    rows_iter = ws.iter_rows(min_row=1, values_only=True)
    header = None
    out = []
    for r in rows_iter:
        if header is None:
            idx = header_index(r)
            if "Tanggal" in idx and "Nama Akun" in idx:
                header = idx
            continue
        tgl = as_date(r[header["Tanggal"]]) if header["Tanggal"] < len(r) else None
        akun = r[header["Nama Akun"]] if header["Nama Akun"] < len(r) else None
        if not tgl or not akun:
            continue
        debet = as_num(r[header["Debet"]]) if "Debet" in header and header["Debet"] < len(r) else 0.0
        kredit = as_num(r[header["Kredit"]]) if "Kredit" in header and header["Kredit"] < len(r) else 0.0
        relasi = r[header["Customer/Vendor"]] if "Customer/Vendor" in header and header["Customer/Vendor"] < len(r) else None
        project = r[header.get("Kategori", -1)] if header.get("Kategori", -1) >= 0 and header["Kategori"] < len(r) else None
        kategori_akun = r[header.get("Kategori Akun", -1)] if header.get("Kategori Akun", -1) >= 0 and header["Kategori Akun"] < len(r) else None
        ket = r[header["Keterangan"]] if "Keterangan" in header and header["Keterangan"] < len(r) else ""
        out.append({
            "tgl": tgl, "akun": str(akun).strip(),
            "project": (project or "").strip() if isinstance(project, str) else "",
            "relasi": (relasi or "").strip() if isinstance(relasi, str) else "",
            "kategori": (kategori_akun or "").strip() if isinstance(kategori_akun, str) else "",
            "ket": (ket or "").strip() if isinstance(ket, str) else "",
            "debet": debet, "kredit": kredit,
        })
    return out


def guess_ledger_name(display_name, known_ledger_names):
    """Mirrors frontend/index.html's guessLedgerName(): picks the longest
    known ledger name (as used on kas/bank/jurnal entries) that appears
    inside the project's display name, so Dashboard totals line up even
    when the two don't match 1:1 (e.g. "Kemhan Bangun Rumdis Cimahi" vs
    "Rumdis Cimahi")."""
    low = (display_name or "").lower()
    best = ""
    for k in known_ledger_names:
        if not k:
            continue
        if k.lower() in low and len(k) > len(best):
            best = k
    return best


def upsert(cur, table, id_col, rows, insert_only_cols=()):
    if not rows:
        return
    cols = list(rows[0].keys())
    col_list = ",".join(f"`{c}`" for c in cols)
    placeholders = ",".join(["%s"] * len(cols))
    update_cols = [c for c in cols if c != id_col and c not in insert_only_cols]
    update_list = ",".join(f"`{c}`=VALUES(`{c}`)" for c in update_cols)
    sql = f"INSERT INTO `{table}` ({col_list}) VALUES ({placeholders})"
    if update_list:
        sql += f" ON DUPLICATE KEY UPDATE {update_list}"
    cur.executemany(sql, [tuple(row[c] for c in cols) for row in rows])


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("xlsx_path")
    ap.add_argument("--dry-run", action="store_true")
    ap.add_argument("--wipe", action="store_true")
    args = ap.parse_args()

    wb = openpyxl.load_workbook(args.xlsx_path, data_only=True)

    coa = load_coa(wb)
    customers, vendors = load_relasi(wb)
    projects = load_projects(wb)
    txns = load_cash_transactions(wb)
    jurnal = load_jurnal(wb)

    known_ledger_names = sorted({t["project"] for t in txns if t["project"]}
                                 | {j["project"] for j in jurnal if j["project"]})
    for p in projects:
        p["ledger_name"] = p["nama"] if p["nama"] in known_ledger_names \
            else guess_ledger_name(p["nama"], known_ledger_names)

    unlinked = [p["nama"] for p in projects if not p["ledger_name"]]

    print(f"COA:        {len(coa)} akun")
    print(f"Customers:  {len(customers)}")
    print(f"Vendors:    {len(vendors)}")
    print(f"Projects:   {len(projects)} ({len(unlinked)} tanpa ledger_name — perlu diisi manual di Master Project)")
    print(f"Transaksi:  {len(txns)} (kas/bank masuk & keluar)")
    print(f"Jurnal:     {len(jurnal)} baris")
    if unlinked:
        print("  Project tanpa link otomatis:", ", ".join(unlinked[:10]) + (" ..." if len(unlinked) > 10 else ""))

    if args.dry_run:
        print("\n--dry-run: tidak ada yang ditulis ke database.")
        return

    host = os.environ.get("MYSQL_HOST")
    db = os.environ.get("MYSQL_DB")
    user = os.environ.get("MYSQL_USER")
    password = os.environ.get("MYSQL_PASS")
    if not all([host, db, user]):
        print("\nERROR: set MYSQL_HOST, MYSQL_DB, MYSQL_USER, MYSQL_PASS env vars.", file=sys.stderr)
        sys.exit(1)

    import pymysql

    conn = pymysql.connect(host=host, user=user, password=password or "", database=db,
                            charset="utf8mb4", autocommit=False)
    try:
        with conn.cursor() as cur:
            if args.wipe:
                cur.execute("SET FOREIGN_KEY_CHECKS=0")
                for t in ["hutang_overrides", "jurnal_umum", "transactions", "projects", "coa", "vendors", "customers"]:
                    cur.execute(f"TRUNCATE TABLE `{t}`")
                cur.execute("SET FOREIGN_KEY_CHECKS=1")

            for i, c in enumerate(coa):
                c["id"] = f"a{i + 1}"
            upsert(cur, "coa", "id", coa)

            for i, c in enumerate(customers):
                c["id"] = f"c{i + 1}"
            for i, v in enumerate(vendors):
                v["id"] = f"v{i + 1}"
            customer_by_name = {c["nama"]: c for c in customers}
            vendor_by_name = {v["nama"]: v for v in vendors}
            upsert(cur, "customers", "id", [
                {"id": c["id"], "kode": c["kode"], "nama": c["nama"], "alamat": c["alamat"], "telp": c["telp"], "email": ""}
                for c in customers])
            upsert(cur, "vendors", "id", [
                {"id": v["id"], "kode": v["kode"], "nama": v["nama"], "alamat": v["alamat"], "telp": v["telp"], "email": ""}
                for v in vendors])

            for i, p in enumerate(projects):
                p["id"] = f"p{i + 1}"
            upsert(cur, "projects", "id", [
                {"id": p["id"], "nama": p["nama"], "ledger_name": p["ledger_name"], "kontrak": p["kontrak"],
                 "rap": p["rap"], "progress": p["progress"], "pemberi_proyek": p["pemberi_proyek"],
                 "cost_center": p["cost_center"], "adm_fee": p["adm_fee"]}
                for p in projects])

            tvals = []
            for i, t in enumerate(txns):
                cust_id = customer_by_name.get(t["relasi"], {}).get("id") if t["jenis"].endswith("masuk") else None
                vend_id = vendor_by_name.get(t["relasi"], {}).get("id") if t["jenis"].endswith("keluar") else None
                tvals.append({
                    "id": f"t{i + 1}", "jenis": t["jenis"], "tgl": t["tgl"], "ref": t["ref"],
                    "akun_kas": t["akun_kas"], "akun_lawan": t["akun_lawan"], "project": t["project"],
                    "relasi": t["relasi"], "customer_id": cust_id, "vendor_id": vend_id,
                    "ket": t["ket"], "debet": t["debet"], "kredit": t["kredit"], "created_by": None,
                })
            upsert(cur, "transactions", "id", tvals, insert_only_cols=("created_by",))

            jvals = [{"id": f"j{i + 1}", "tgl": j["tgl"], "ref": f"JU-{i + 1:04d}", "akun": j["akun"],
                      "project": j["project"], "relasi": j["relasi"], "kategori": j["kategori"],
                      "ket": j["ket"], "debet": j["debet"], "kredit": j["kredit"], "created_by": None}
                     for i, j in enumerate(jurnal)]
            upsert(cur, "jurnal_umum", "id", jvals, insert_only_cols=("created_by",))

        conn.commit()
        print("\nImport selesai dan sudah di-commit ke database.")
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()


if __name__ == "__main__":
    main()
