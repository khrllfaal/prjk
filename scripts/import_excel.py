#!/usr/bin/env python3
"""
Import the company's Excel workbook (KAS BESAR/KECIL, BANK*, COA,
MASTER CUST & VENDOR, Nama Project, JURNAL UMUM) into the Supabase
Postgres database created by backend/supabase/migrations/0001_init.sql.

This script is intentionally NOT wired to any hardcoded file path or
committed data — it takes the xlsx path and DB connection as arguments/
env vars so the real financial data never has to touch git.

Usage:
    export DATABASE_URL="postgresql://postgres:<password>@<host>:5432/postgres"
    python3 scripts/import_excel.py /path/to/LAPORAN_KEUANGAN.xlsx

Options:
    --dry-run   parse and print row counts only, no DB writes
    --wipe      truncate the target tables before importing (fresh import)
"""
import argparse
import os
import re
import sys
from datetime import date, datetime

import openpyxl

CASH_SHEETS = {
    "KAS BESAR": "Kas Besar",
    "KAS KECIL": "Kas Kecil",
    "BANK BCA": "Bank BCA",
    "BANK BJB": "Bank BJB",
    "BANK BRI": "Bank BRI",
    "BANK BNI": "Bank BNI",
    "MANDIRI": "Bank MANDIRI",
}


def header_index(row):
    """Map non-empty header cell text -> column index (0-based)."""
    idx = {}
    for i, v in enumerate(row):
        if v:
            idx[str(v).strip()] = i
    return idx


def as_date(v):
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    if isinstance(v, str) and v.strip():
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
            try:
                return datetime.strptime(v.strip(), fmt).date()
            except ValueError:
                pass
    return None


def as_num(v):
    if v is None or v == "":
        return 0.0
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0.0


def load_coa(wb):
    ws = wb["COA"]
    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        kode = r[0]
        if not kode or not re.match(r"^\d+(\.\d+)*$", str(kode)):
            continue
        nama = (r[1] or "").strip()
        level = str(kode).count(".") + 1
        head = str(kode).split(".")[0]
        tipe = {"1": "asset", "2": "liability", "3": "equity", "4": "revenue",
                "5": "cogs", "6": "expense", "7": "other", "8": "tax"}.get(head, "other")
        rows.append({"kode": str(kode), "nama": nama, "level": level, "tipe": tipe})
    return rows


def load_relasi(wb):
    ws = wb["MASTER CUST & VENDOR"]
    customers, vendors = [], []
    ci = vi = 1
    for r in ws.iter_rows(min_row=3, values_only=True):
        nama = r[1] if len(r) > 1 else None
        if not nama or not isinstance(nama, str) or nama.startswith("Note"):
            continue
        rec = {"nama": nama.strip(), "alamat": (r[2] or "").strip() if len(r) > 2 else "",
               "telp": str(r[3]).strip() if len(r) > 3 and r[3] else ""}
        if re.match(r"^Pr-", nama, re.I):
            rec["kode"] = f"CU{ci:03d}"
            customers.append(rec)
            ci += 1
        else:
            rec["kode"] = f"VE{vi:03d}"
            vendors.append(rec)
            vi += 1
    return customers, vendors


PEMBERI_PREFIXES = ["Kemhan", "Bina Marga", "PUPR", "Korem", "Kodim", "Koramil", "BJB", "Dinas", "Diknas"]


def guess_pemberi(nama):
    for p in PEMBERI_PREFIXES:
        if nama.lower().startswith(p.lower()) or f" {p.lower()} " in f" {nama.lower()} ":
            return p
    return ""


def load_projects(wb):
    ws = wb["Nama Project"]
    rows = []
    for r in ws.iter_rows(min_row=3, values_only=True):
        nama = r[1] if len(r) > 1 else None
        if not nama or not isinstance(nama, str):
            continue
        rows.append({"nama": nama.strip(), "kontrak": 0, "rap": 0, "progress": 0,
                     "pemberi_proyek": guess_pemberi(nama)})
    return rows


def load_cash_transactions(wb, project_names, customer_names, vendor_names):
    txns = []
    counters = {}
    for sheet, akun_kas in CASH_SHEETS.items():
        if sheet not in wb.sheetnames:
            continue
        ws = wb[sheet]
        rows_iter = ws.iter_rows(min_row=1, values_only=True)
        header = header_index(next(rows_iter))
        is_bank = akun_kas.lower().startswith("bank")
        col_tgl = header.get("Tanggal")
        col_ket = header.get("Keterangan")
        col_relasi = header.get("Customer/Vendor")
        col_project = header.get("Nama Project", header.get("Kategori Produk"))
        col_akun = header.get("Nama Akun")
        col_kategori = header.get("Kategori Akun")
        col_debet = header.get("Debet")
        col_kredit = header.get("Kredit")
        if col_tgl is None or col_akun is None:
            continue
        for r in rows_iter:
            tgl = as_date(r[col_tgl]) if col_tgl < len(r) else None
            akun_lawan = r[col_akun] if col_akun < len(r) else None
            if not tgl or not akun_lawan:
                continue
            debet = as_num(r[col_debet]) if col_debet is not None and col_debet < len(r) else 0.0
            kredit = as_num(r[col_kredit]) if col_kredit is not None and col_kredit < len(r) else 0.0
            if debet == 0 and kredit == 0:
                continue
            is_in = debet > 0
            jenis = ("bank_" if is_bank else "kas_") + ("masuk" if is_in else "keluar")
            pref = ("BI" if is_bank else "CI") if is_in else ("BO" if is_bank else "CO")
            mo = tgl.strftime("%y%m")
            counters[pref] = counters.get(pref, 0) + 1
            relasi = r[col_relasi] if col_relasi is not None and col_relasi < len(r) else None
            relasi = (relasi or "").strip() if isinstance(relasi, str) else ""
            project = r[col_project] if col_project is not None and col_project < len(r) else None
            project = (project or "").strip() if isinstance(project, str) else ""
            ket = r[col_ket] if col_ket is not None and col_ket < len(r) else ""
            kategori = r[col_kategori] if col_kategori is not None and col_kategori < len(r) else ""
            txns.append({
                "tgl": tgl, "ref": f"{pref}-{mo}-{counters[pref]:04d}",
                "akun_kas": akun_kas, "akun_lawan": str(akun_lawan).strip(),
                "project": project, "relasi": relasi,
                "customer_id": None, "vendor_id": None,  # resolved by caller
                "kategori": str(kategori).strip() if kategori else "",
                "ket": str(ket).strip() if ket else "",
                "debet": debet, "kredit": kredit, "jenis": jenis,
            })
    return txns


def load_jurnal(wb):
    if "JURNAL UMUM" not in wb.sheetnames:
        return []
    ws = wb["JURNAL UMUM"]
    rows_iter = ws.iter_rows(min_row=1, values_only=True)
    header = None
    out = []
    for r in rows_iter:
        if header is None:
            idx = header_index(r)
            if "Tanggal" in idx and "Nama Akun" in idx:
                header = idx
            continue
        tgl = as_date(r[header["Tanggal"]]) if header["Tanggal"] < len(r) else None
        akun = r[header["Nama Akun"]] if header["Nama Akun"] < len(r) else None
        if not tgl or not akun:
            continue
        debet = as_num(r[header["Debet"]]) if "Debet" in header and header["Debet"] < len(r) else 0.0
        kredit = as_num(r[header["Kredit"]]) if "Kredit" in header and header["Kredit"] < len(r) else 0.0
        relasi = r[header["Customer/Vendor"]] if "Customer/Vendor" in header and header["Customer/Vendor"] < len(r) else None
        project = r[header.get("Kategori", -1)] if header.get("Kategori", -1) >= 0 and header["Kategori"] < len(r) else None
        kategori_akun = r[header.get("Kategori Akun", -1)] if header.get("Kategori Akun", -1) >= 0 and header["Kategori Akun"] < len(r) else None
        ket = r[header["Keterangan"]] if "Keterangan" in header and header["Keterangan"] < len(r) else ""
        out.append({
            "tgl": tgl, "akun": str(akun).strip(),
            "project": (project or "").strip() if isinstance(project, str) else "",
            "relasi": (relasi or "").strip() if isinstance(relasi, str) else "",
            "kategori": (kategori_akun or "").strip() if isinstance(kategori_akun, str) else "",
            "ket": (ket or "").strip() if isinstance(ket, str) else "",
            "debet": debet, "kredit": kredit,
        })
    return out


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("xlsx_path")
    ap.add_argument("--dry-run", action="store_true")
    ap.add_argument("--wipe", action="store_true")
    args = ap.parse_args()

    wb = openpyxl.load_workbook(args.xlsx_path, data_only=True)

    coa = load_coa(wb)
    customers, vendors = load_relasi(wb)
    projects = load_projects(wb)
    project_names = {p["nama"] for p in projects}
    customer_by_name = {c["nama"]: c for c in customers}
    vendor_by_name = {v["nama"]: v for v in vendors}
    txns = load_cash_transactions(wb, project_names, customer_by_name, vendor_by_name)
    jurnal = load_jurnal(wb)

    print(f"COA:        {len(coa)} akun")
    print(f"Customers:  {len(customers)}")
    print(f"Vendors:    {len(vendors)}")
    print(f"Projects:   {len(projects)}")
    print(f"Transaksi:  {len(txns)} (kas/bank masuk & keluar)")
    print(f"Jurnal:     {len(jurnal)} baris")

    if args.dry_run:
        print("\n--dry-run: tidak ada yang ditulis ke database.")
        return

    dsn = os.environ.get("DATABASE_URL")
    if not dsn:
        print("\nERROR: set DATABASE_URL env var (Supabase Settings > Database > Connection string).",
              file=sys.stderr)
        sys.exit(1)

    import psycopg2
    from psycopg2.extras import execute_values

    conn = psycopg2.connect(dsn)
    conn.autocommit = False
    try:
        with conn.cursor() as cur:
            if args.wipe:
                cur.execute("truncate transactions, jurnal_umum, hutang_dagang, projects, coa, vendors, customers restart identity cascade")

            # --- COA ---
            execute_values(cur,
                "insert into coa (id, kode, nama, level, tipe) values %s on conflict (id) do nothing",
                [(f"a{i+1}", c["kode"], c["nama"], c["level"], c["tipe"]) for i, c in enumerate(coa)])

            # --- customers / vendors ---
            for i, c in enumerate(customers):
                c["id"] = f"c{i+1}"
            for i, v in enumerate(vendors):
                v["id"] = f"v{i+1}"
            execute_values(cur,
                "insert into customers (id, kode, nama, alamat, telp, email) values %s on conflict (id) do nothing",
                [(c["id"], c["kode"], c["nama"], c["alamat"], c["telp"], "") for c in customers])
            execute_values(cur,
                "insert into vendors (id, kode, nama, alamat, telp, email) values %s on conflict (id) do nothing",
                [(v["id"], v["kode"], v["nama"], v["alamat"], v["telp"], "") for v in vendors])

            # --- projects ---
            for i, p in enumerate(projects):
                p["id"] = f"p{i+1}"
            execute_values(cur,
                "insert into projects (id, nama, kontrak, rap, progress, pemberi_proyek) values %s on conflict (id) do nothing",
                [(p["id"], p["nama"], p["kontrak"], p["rap"], p["progress"], p["pemberi_proyek"]) for p in projects])

            # --- transactions (resolve customer_id / vendor_id from relasi name) ---
            tvals = []
            for i, t in enumerate(txns):
                cust_id = customer_by_name.get(t["relasi"], {}).get("id") if t["jenis"].endswith("masuk") else None
                vend_id = vendor_by_name.get(t["relasi"], {}).get("id") if t["jenis"].endswith("keluar") else None
                tvals.append((
                    f"t{i+1}", t["jenis"], t["tgl"], t["ref"], t["akun_kas"], t["akun_lawan"],
                    t["project"], t["relasi"], cust_id, vend_id, t["ket"], t["debet"], t["kredit"],
                ))
            execute_values(cur,
                """insert into transactions
                   (id, jenis, tgl, ref, akun_kas, akun_lawan, project, relasi, customer_id, vendor_id, ket, debet, kredit)
                   values %s on conflict (id) do nothing""",
                tvals, page_size=500)

            # --- jurnal umum ---
            jvals = [(f"j{i+1}", j["tgl"], f"JU-{i+1:04d}", j["akun"], j["project"], j["relasi"],
                      j["kategori"], j["ket"], j["debet"], j["kredit"]) for i, j in enumerate(jurnal)]
            execute_values(cur,
                """insert into jurnal_umum (id, tgl, ref, akun, project, relasi, kategori, ket, debet, kredit)
                   values %s on conflict (id) do nothing""",
                jvals)

        conn.commit()
        print("\nImport selesai dan sudah di-commit ke database.")
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()


if __name__ == "__main__":
    main()
